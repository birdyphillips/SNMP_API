import requests
import json
import os
from datetime import datetime
from dotenv import load_dotenv

# Load credentials from .env file
load_dotenv()

# Excel export functionality
try:
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# Latency bin constants
# ---------------------------------------------------------------------------
US_BIN_EDGES_MS = [
    0, 0.05, 0.10, 0.25, 0.50, 1.00, 2.00, 5.00, 10.00,
    20.00, 30.00, 40.00, 50.00, 100.00, 150.00, 200.00, 500.00,
]
DS_BIN_EDGES_MS = [
    0, 0.05, 0.10, 0.25, 0.50, 1.00, 2.00, 5.00, 10.00,
    20.00, 30.00, 40.00, 50.00, 60.00, 70.00, 80.00, 500.00,
]
NUM_BINS = 16


# ---------------------------------------------------------------------------
# Percentile & average calculations
# ---------------------------------------------------------------------------

def _get_edges(bin_edges):
    """Return bin_edges if provided, else default to US edges."""
    return bin_edges if bin_edges is not None else US_BIN_EDGES_MS


def calc_percentile(deltas, percentile, bin_edges=None):
    """Linear interpolation percentile from bin deltas."""
    edges = _get_edges(bin_edges)
    total = sum(deltas)
    if total == 0:
        return 0.0
    target = total * percentile
    cumulative = 0
    for i, count in enumerate(deltas):
        cumulative += count
        if cumulative >= target:
            prev_cum = cumulative - count
            bin_low  = edges[i]
            bin_high = edges[i + 1]
            denom = count if count > 0 else 1
            return bin_low + ((target - prev_cum) / denom) * (bin_high - bin_low)
    return edges[-1]


def calc_weighted_avg(deltas, bin_edges=None):
    """Weighted average latency: sum(delta_i x bin_avg_i) / total."""
    edges = _get_edges(bin_edges)
    total = sum(deltas)
    if total == 0:
        return 0.0
    weighted = sum(
        deltas[i] * (edges[i] + edges[i + 1]) / 2
        for i in range(NUM_BINS)
    )
    return weighted / total


def calc_percentile_avg(deltas, percentile, bin_edges=None):
    """AVG method: return the bin midpoint of the first bin where
    cumulative count >= percentile target."""
    edges = _get_edges(bin_edges)
    total = sum(deltas)
    if total == 0:
        return 0.0
    target = total * percentile
    cumulative = 0
    for i, count in enumerate(deltas):
        cumulative += count
        if cumulative >= target:
            return (edges[i] + edges[i + 1]) / 2
    return (edges[-2] + edges[-1]) / 2


def calc_latency_stats(bins, bin_edges=None):
    """Return all latency metrics for a bin list in one call."""
    return {
        "total":        sum(bins),
        "weighted_avg": calc_weighted_avg(bins, bin_edges),
        "p50":          calc_percentile(bins, 0.50,  bin_edges),
        "p99":          calc_percentile(bins, 0.99,  bin_edges),
        "p999":         calc_percentile(bins, 0.999, bin_edges),
        "p50a":         calc_percentile_avg(bins, 0.50,  bin_edges),
        "p99a":         calc_percentile_avg(bins, 0.99,  bin_edges),
        "p999a":        calc_percentile_avg(bins, 0.999, bin_edges),
    }


# ---------------------------------------------------------------------------
# SNMP text file parser — pure Python, no regex
# ---------------------------------------------------------------------------

def _find_between(text, start, end):
    s = text.find(start)
    if s == -1:
        return ""
    s += len(start)
    e = text.find(end, s)
    return text[s:e] if e != -1 else text[s:]


def _split_oid_line(line):
    """
    Parse a raw SNMP OID line:
      SNMPv2-SMI::enterprises.<oid_tail> = <type>: <value>
    Returns (oid_tail, type_str, value_str) or None.
    """
    eq = line.find(" = ")
    if eq == -1:
        return None
    lhs = line[:eq].strip()
    rhs = line[eq + 3:].strip()
    colon = rhs.find(": ")
    if colon == -1:
        return None
    type_str  = rhs[:colon].strip()
    value_str = rhs[colon + 2:].strip()
    for prefix in ("SNMPv2-SMI::", "iso."):
        if lhs.startswith(prefix):
            lhs = lhs[len(prefix):]
    return lhs, type_str, value_str


def _is_integer(s):
    try:
        int(s)
        return True
    except (ValueError, TypeError):
        return False


def _extract_section(content, section_title):
    """Extract lines under a section header followed by a === divider."""
    lines = content.splitlines()
    result = []
    in_section = False
    found_divider = False
    for line in lines:
        if not in_section:
            if line.strip() == section_title.strip():
                in_section = True
            continue
        if not found_divider:
            if line.startswith("="):
                found_divider = True
            continue
        if line.strip() == "":
            break
        result.append(line)
    return result


def parse_snmp_timestamp(filepath):
    """Extract collection timestamp from SNMP file header. Returns dict or None."""
    with open(filepath, "r") as f:
        for line in f:
            prefix = "SNMP Collection - "
            if line.startswith(prefix):
                ts = line[len(prefix):].strip()
                if len(ts) >= 19:
                    try:
                        return {
                            "year":   int(ts[0:4]),
                            "month":  int(ts[5:7]),
                            "day":    int(ts[8:10]),
                            "hour":   int(ts[11:13]),
                            "minute": int(ts[14:16]),
                            "second": int(ts[17:19]),
                            "raw":    ts[:19],
                        }
                    except ValueError:
                        pass
    return None


def ts_to_seconds(ts):
    """Convert timestamp dict to approximate total seconds (for delta calculations)."""
    if ts is None:
        return None
    y, mo, d = ts["year"], ts["month"], ts["day"]
    days = y * 365 + y // 4 - y // 100 + y // 400
    month_days = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    days += month_days[mo - 1] + d
    leap = (y % 4 == 0 and y % 100 != 0) or (y % 400 == 0)
    if leap and mo > 2:
        days += 1
    return days * 86400 + ts["hour"] * 3600 + ts["minute"] * 60 + ts["second"]


def ts_delta_seconds(ts_before, ts_after):
    """Return seconds between two timestamp dicts."""
    a = ts_to_seconds(ts_before)
    b = ts_to_seconds(ts_after)
    if a is None or b is None:
        return None
    return b - a


def parse_flow_stats(filepath):
    """Parse Flow Stats Table. Returns {sfid: {packets, octets, dropped}}."""
    with open(filepath, "r") as f:
        content = f.read()
    section_lines = _extract_section(content, "Flow Stats Table (Entry Qos Service Flow Octets)")
    if not section_lines:
        section_lines = _extract_section(content, "Flow Stats Table")
    sub_map = {"1": "packets", "2": "octets", "8": "dropped"}
    stats = {}
    for line in section_lines:
        parsed = _split_oid_line(line)
        if not parsed:
            continue
        oid, type_str, val_str = parsed
        if type_str not in ("Counter64", "Counter32") or not _is_integer(val_str):
            continue
        parts = oid.split(".")
        for i in range(len(parts) - 4):
            if parts[i] == "4" and parts[i+1] == "1" and parts[i+3] == "2":
                sub, sfid_str = parts[i+2], parts[i+4]
                if sub in sub_map and _is_integer(sfid_str):
                    sfid = int(sfid_str)
                    stats.setdefault(sfid, {"packets": 0, "octets": 0, "dropped": 0})
                    stats[sfid][sub_map[sub]] = int(val_str)
                break
    return stats


def parse_latency_bins(filepath):
    """Parse Latency Stats Table. Returns {sfid: {sub_oid: count}}."""
    with open(filepath, "r") as f:
        content = f.read()
    section_lines = _extract_section(content, "Latency Stats Table")
    bins = {}
    for line in section_lines:
        parsed = _split_oid_line(line)
        if not parsed:
            continue
        oid, type_str, val_str = parsed
        if type_str not in ("Counter64", "Gauge32") or not _is_integer(val_str):
            continue
        parts = oid.split(".")
        for i in range(len(parts) - 5):
            if parts[i] == "29" and parts[i+1] == "2" and parts[i+2] == "1" and parts[i+4] == "2":
                sub_oid_str, sfid_str = parts[i+3], parts[i+5]
                if _is_integer(sub_oid_str) and _is_integer(sfid_str):
                    bins.setdefault(int(sfid_str), {})[int(sub_oid_str)] = int(val_str)
                break
    return bins


def parse_congestion_stats(filepath):
    """Parse Congestion Stats Table. Returns {sfid: {aqm_drops, congestion_marked, sanctioned}}."""
    with open(filepath, "r") as f:
        content = f.read()
    section_lines = _extract_section(content, "Congestion Stats Table")
    sub_map = {"1": "aqm_drops", "3": "congestion_marked", "4": "sanctioned"}
    stats = {}
    for line in section_lines:
        parsed = _split_oid_line(line)
        if not parsed:
            continue
        oid, type_str, val_str = parsed
        if type_str != "Counter64" or not _is_integer(val_str):
            continue
        parts = oid.split(".")
        for i in range(len(parts) - 4):
            if parts[i] == "30" and parts[i+1] == "1" and parts[i+3] == "2":
                sub, sfid_str = parts[i+2], parts[i+4]
                if sub in sub_map and _is_integer(sfid_str):
                    sfid = int(sfid_str)
                    stats.setdefault(sfid, {"aqm_drops": 0, "congestion_marked": 0, "sanctioned": 0})
                    stats[sfid][sub_map[sub]] = int(val_str)
                break
    return stats


def compute_deltas(before_bins, after_bins):
    """Compute per-SFID bin deltas (after - before, floored at 0)."""
    results = {}
    for sfid in sorted(set(before_bins) & set(after_bins)):
        before_vals, after_vals, deltas = [], [], []
        for sub in range(3, 3 + NUM_BINS):
            bv = before_bins[sfid].get(sub, 0)
            av = after_bins[sfid].get(sub, 0)
            before_vals.append(bv)
            after_vals.append(av)
            deltas.append(max(av - bv, 0))
        if sum(deltas) > 0:
            results[sfid] = {"before": before_vals, "after": after_vals, "deltas": deltas}
    return results


def compute_flow_deltas(fs_before, fs_after):
    """Compute per-SFID flow stat deltas."""
    results = {}
    for sfid in sorted(set(fs_before) & set(fs_after)):
        d_octets  = max(fs_after[sfid]["octets"]  - fs_before[sfid]["octets"],  0)
        d_packets = max(fs_after[sfid]["packets"] - fs_before[sfid]["packets"], 0)
        d_dropped = max(fs_after[sfid]["dropped"] - fs_before[sfid]["dropped"], 0)
        if d_packets == 0 and d_octets == 0:
            continue
        results[sfid] = {"d_packets": d_packets, "d_octets": d_octets, "d_dropped": d_dropped}
    return results


def compute_throughput_and_loss(before_file, after_file, duration_s=None):
    """Compute per-SFID throughput (Mbps) and packet loss from flow stat deltas."""
    fs_before = parse_flow_stats(before_file)
    fs_after  = parse_flow_stats(after_file)
    if not fs_before or not fs_after:
        return {}
    if duration_s is None:
        duration_s = ts_delta_seconds(parse_snmp_timestamp(before_file), parse_snmp_timestamp(after_file))
    if not duration_s or duration_s <= 0:
        return {}
    results = {}
    for sfid, d in compute_flow_deltas(fs_before, fs_after).items():
        total_pkts = d["d_packets"] + d["d_dropped"]
        results[sfid] = {
            "throughput_mbps": (d["d_octets"] * 8) / (duration_s * 1_000_000),
            "lost_packets":    d["d_dropped"],
            "total_packets":   total_pkts,
            "loss_pct":        (d["d_dropped"] / total_pkts * 100) if total_pkts > 0 else 0.0,
        }
    return results


def parse_snmp_pair(before_file, after_file, direction="US"):
    """Parse a before/after SNMP file pair and return a complete results dict."""
    edges = US_BIN_EDGES_MS if direction == "US" else DS_BIN_EDGES_MS
    before_bins = parse_latency_bins(before_file)
    after_bins  = parse_latency_bins(after_file)
    lat_deltas  = compute_deltas(before_bins, after_bins)
    for sfid, d in lat_deltas.items():
        d["stats"] = calc_latency_stats(d["deltas"], edges)
    tp = compute_throughput_and_loss(before_file, after_file)
    cong_before = parse_congestion_stats(before_file)
    cong_after  = parse_congestion_stats(after_file)
    congestion  = {}
    for sfid in sorted(set(cong_before) & set(cong_after)):
        cb, ca = cong_before[sfid], cong_after[sfid]
        congestion[sfid] = {
            "aqm_drops_delta":         max(ca["aqm_drops"]         - cb["aqm_drops"],         0),
            "congestion_marked_delta": max(ca["congestion_marked"] - cb["congestion_marked"], 0),
            "sanctioned_delta":        max(ca["sanctioned"]        - cb["sanctioned"],        0),
        }
    return {
        "direction":      direction,
        "bin_edges":      edges,
        "latency_deltas": lat_deltas,
        "throughput":     tp,
        "congestion":     congestion,
        "duration_s":     ts_delta_seconds(parse_snmp_timestamp(before_file), parse_snmp_timestamp(after_file)),
    }


# ---------------------------------------------------------------------------
# Data extraction from Scope API response
# ---------------------------------------------------------------------------

def _sf_key_to_sfid(key):
    """Extract SFID int from keys like '2.93890' -> 93890."""
    parts = key.split(".")
    if len(parts) >= 2:
        try:
            return int(parts[-1])
        except ValueError:
            pass
    return None


def extract_service_flow_data(data):
    """
    Parse Scope API JSON response into per-service-flow structures.
    API returns each field as a dict keyed by '2.<sfid>', e.g.:
      "qos_service_flow_direction": {"2.93890": "upstream", "2.93891": "downstream"}
    """
    service_flows = {}

    if not data or not data.get("data"):
        return service_flows

    for item in data["data"]:
        # Discover all SFIDs from qos_service_flow_direction keys
        direction_map = item.get("qos_service_flow_direction", {})
        if not isinstance(direction_map, dict):
            continue

        for sf_key, direction in direction_map.items():
            sfid = _sf_key_to_sfid(sf_key)
            if sfid is None:
                continue

            def _get(field, default=0):
                val = item.get(field, {})
                if isinstance(val, dict):
                    return val.get(sf_key, default)
                return val if val is not None else default

            def _get_int(field, default=0):
                try:
                    return int(_get(field, default) or default)
                except (ValueError, TypeError):
                    return default

            # Service class: keyed by '2.1.<sfid>'
            sc_key = f"2.1.{sfid}"
            sc_map = item.get("qos_param_set_service_class_name", {})
            service_class = sc_map.get(sc_key, "") if isinstance(sc_map, dict) else ""

            # Max traffic rate: keyed by '2.<sfid>.1'
            rate_key = f"{sf_key}.1"
            rate_map = item.get("qos_param_set_max_traffic_rate", {})
            max_rate = int(rate_map.get(rate_key, 0) or 0) if isinstance(rate_map, dict) else 0

            # Max traffic burst: keyed by '2.<sfid>.1'
            burst_map = item.get("qos_param_set_max_traffic_burst", {})
            max_burst = int(burst_map.get(rate_key, 0) or 0) if isinstance(burst_map, dict) else 0

            # Max concat burst: keyed by '2.<sfid>.1'
            concat_map = item.get("qos_param_set_max_concat_burst", {})
            max_concat = int(concat_map.get(rate_key, 0) or 0) if isinstance(concat_map, dict) else 0

            # Priority: keyed by '2.<sfid>.1'
            priority_map = item.get("qos_param_set_priority", {})
            priority = int(priority_map.get(rate_key, 0) or 0) if isinstance(priority_map, dict) else 0

            # Latency bins 1-16
            bins = []
            for b in range(1, NUM_BINS + 1):
                bin_map = item.get(f"docsQosSfLatencyBin{b}Pkts", {})
                val = bin_map.get(sf_key, 0) if isinstance(bin_map, dict) else 0
                bins.append(int(val) if val else 0)

            dir_str = direction.lower() if isinstance(direction, str) else str(direction)

            service_flows[sfid] = {
                "sf_key":              sf_key,
                "direction":           dir_str,
                "primary":             _get("qos_service_flow_primary", ""),
                "service_class":       service_class,
                "max_traffic_rate":    max_rate,
                "max_traffic_burst":   max_burst,
                "max_concat_burst":    max_concat,
                "priority":            priority,
                "packets":             _get_int("qos_service_flow_pkts"),
                "octets":              _get_int("qos_service_flow_octets"),
                "policed_drop_pkts":   _get_int("qos_service_flow_policed_drop_kts"),
                "policed_delay_pkts":  _get_int("qos_service_flow_policed_delay_pkts"),
                "aqm_dropped_pkts":    _get_int("qos_service_flow_aqm_dropped_pkts"),
                "buffer_size":         _get_int("qos_service_flow_buffer_size"),
                "max_latency":         _get_int("docsQosSfLatencyMaxLatency"),
                "hist_updates":        _get_int("docsQosSfLatencyNumHistUpdates"),
                "latency_bins":        bins,
                "congestion_sanctioned":  _get_int("docsQosSfCongestionSanctionedPkts"),
                "congestion_ect0":        _get_int("docsQosSfCongestionTotalEct0Pkts"),
                "congestion_ect1":        _get_int("docsQosSfCongestionTotalEct1Pkts"),
                "congestion_ce_marked":   _get_int("docsQosSfCongestionCeMarkedEct1Pkts"),
                "congestion_arrived_ce":  _get_int("docsQosSfCongestionArrivedCePkts"),
                "tx_retries":   _get_int("qos_service_us_stats_tx_retries"),
                "tx_exceededs": _get_int("qos_service_us_stats_tx_exceededs"),
                "rq_retries":   _get_int("qos_service_us_stats_rq_retries"),
                "rq_exceededs": _get_int("qos_service_us_stats_rq_exceededs"),
            }

    return service_flows


# ---------------------------------------------------------------------------
# Excel styling constants
# ---------------------------------------------------------------------------

if OPENPYXL_AVAILABLE:
    _HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    _HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    _CALC_FILL = PatternFill("solid", fgColor="D9E2F3")
    _RESULT_FILL = PatternFill("solid", fgColor="C6EFCE")
    _INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
    _THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    _CENTER = Alignment(horizontal="center", vertical="center")
    _BOLD = Font(bold=True, size=11)


def _styled_cell(ws, row, col, value, font=None, fill=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = _CENTER
    cell.border = _THIN_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# ---------------------------------------------------------------------------
# Excel sheet writers
# ---------------------------------------------------------------------------

def write_bin_edges_sheet(wb, bin_edges, title="Bin_Edges", direction="US"):
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = "FFC000" if direction == "US" else "FF6600"

    ws.merge_cells("A1:C1")
    ws["A1"] = f"EDITABLE {direction} BIN EDGES (ms) — Change values below to recalculate all sheets"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = _CENTER

    _styled_cell(ws, 3, 1, "Bin #", font=_HEADER_FONT, fill=_HEADER_FILL)
    _styled_cell(ws, 3, 2, "Lower Edge (ms)", font=_HEADER_FONT, fill=_HEADER_FILL)
    _styled_cell(ws, 3, 3, "Upper Edge (ms)", font=_HEADER_FONT, fill=_HEADER_FILL)

    for i in range(NUM_BINS):
        row = 4 + i
        _styled_cell(ws, row, 1, i + 1)
        _styled_cell(ws, row, 2, bin_edges[i], fill=_INPUT_FILL, fmt="0.00")
        _styled_cell(ws, row, 3, bin_edges[i + 1], fill=_INPUT_FILL, fmt="0.00")

    for i, w in enumerate([8, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sf_sheet(wb, sheet_name, sf_data, bin_edges, direction="US"):
    bins = sf_data["latency_bins"]
    ws = wb.create_sheet(title=sheet_name)
    edges_sheet = f"Bin_Edges_{direction}"

    dir_label = "UPSTREAM" if direction == "US" else "DOWNSTREAM"
    ws.merge_cells("A1:J1")
    ws["A1"] = f"SCOPE API {dir_label} LATENCY — {sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    # Metadata row
    _styled_cell(ws, 2, 1, "Service Class:", font=_BOLD)
    _styled_cell(ws, 2, 2, sf_data.get("service_class", ""), fill=_RESULT_FILL)
    _styled_cell(ws, 2, 4, "Packets:", font=_BOLD)
    _styled_cell(ws, 2, 5, sf_data["packets"], fill=_RESULT_FILL)
    _styled_cell(ws, 2, 6, "Octets:", font=_BOLD)
    _styled_cell(ws, 2, 7, sf_data["octets"], fill=_RESULT_FILL)

    headers = [
        "BIN", "LOWER (ms)", "UPPER (ms)", "AVG (ms)",
        "COUNT", "CUMULATIVE", "CUMULATIVE %", "BIN %",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 3, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    stats = calc_latency_stats(bins, bin_edges)
    total = stats["total"]
    cumulative = 0
    for i in range(NUM_BINS):
        row = 4 + i
        be_row = 4 + i
        cumulative += bins[i]
        cum_pct = (cumulative / total * 100) if total > 0 else 0
        bin_pct = (bins[i] / total * 100) if total > 0 else 0

        _styled_cell(ws, row, 1, i + 1)
        c = _styled_cell(ws, row, 2, None, fmt="0.00")
        c.value = f"='{edges_sheet}'!B{be_row}"
        c = _styled_cell(ws, row, 3, None, fmt="0.00")
        c.value = f"='{edges_sheet}'!C{be_row}"
        c = _styled_cell(ws, row, 4, None, fill=_CALC_FILL, fmt="0.0000")
        c.value = f"=(B{row}+C{row})/2"
        _styled_cell(ws, row, 5, bins[i], fill=_INPUT_FILL)
        _styled_cell(ws, row, 6, cumulative, fill=_CALC_FILL)
        _styled_cell(ws, row, 7, round(cum_pct, 2), fill=_CALC_FILL, fmt="0.00")
        _styled_cell(ws, row, 8, round(bin_pct, 2), fill=_CALC_FILL, fmt="0.00")

    total_row = 4 + NUM_BINS
    _styled_cell(ws, total_row, 1, "TOTAL", font=_BOLD)
    _styled_cell(ws, total_row, 5, total, font=_BOLD, fill=_INPUT_FILL)

    # --- Percentile results (Linear Interpolation) ---
    pct_row = total_row + 2
    ws.merge_cells(f"A{pct_row}:H{pct_row}")
    ws.cell(row=pct_row, column=1, value="PERCENTILE RESULTS (LINEAR INTERPOLATION)").font = Font(bold=True, size=12)

    for label, pct_val, result in [
        ("P50",   0.50,  stats["p50"]),
        ("P99",   0.99,  stats["p99"]),
        ("P99.9", 0.999, stats["p999"]),
    ]:
        pct_row += 1
        _styled_cell(ws, pct_row, 1, label, font=_BOLD)
        _styled_cell(ws, pct_row, 2, round(result, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, pct_row, 3, "ms", font=_BOLD)
        _styled_cell(ws, pct_row, 4, f"target={round(total * pct_val, 1)} pkts", fill=_CALC_FILL)

    # --- Percentile results (AVG Method) ---
    avg_row = pct_row + 2
    ws.merge_cells(f"A{avg_row}:H{avg_row}")
    ws.cell(row=avg_row, column=1, value="PERCENTILE RESULTS (AVG METHOD)").font = Font(bold=True, size=12)

    for label, result in [
        ("P50 AVG",   stats["p50a"]),
        ("P99 AVG",   stats["p99a"]),
        ("P99.9 AVG", stats["p999a"]),
    ]:
        avg_row += 1
        _styled_cell(ws, avg_row, 1, label, font=_BOLD)
        _styled_cell(ws, avg_row, 2, round(result, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, avg_row, 3, "ms", font=_BOLD)

    # --- Weighted average ---
    w_row = avg_row + 2
    _styled_cell(ws, w_row, 1, "Weighted Avg", font=_BOLD)
    _styled_cell(ws, w_row, 2, round(stats["weighted_avg"], 4), fill=_RESULT_FILL, fmt="0.0000")
    _styled_cell(ws, w_row, 3, "ms", font=_BOLD)

    # --- Max latency & hist updates ---
    w_row += 1
    _styled_cell(ws, w_row, 1, "Max Latency", font=_BOLD)
    _styled_cell(ws, w_row, 2, sf_data.get("max_latency", 0), fill=_CALC_FILL)
    _styled_cell(ws, w_row, 3, "ms", font=_BOLD)
    _styled_cell(ws, w_row, 4, "Hist Updates:", font=_BOLD)
    _styled_cell(ws, w_row, 5, sf_data.get("hist_updates", 0), fill=_CALC_FILL)

    for i, w in enumerate([14, 14, 14, 14, 14, 16, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(wb, service_flows, bin_edges, mac, timestamp):
    ws = wb.create_sheet(title="Summary")
    ws.sheet_properties.tabColor = "4472C4"

    ws.merge_cells("A1:P1")
    ws["A1"] = f"SCOPE API LATENCY SUMMARY — {mac}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    _styled_cell(ws, 2, 1, "Collected:", font=_BOLD)
    _styled_cell(ws, 2, 2, timestamp)

    headers = [
        "SID", "Direction", "Service Class", "Total Bin Pkts",
        "Weighted Avg (ms)", "P50 (ms)", "P99 (ms)", "P99.9 (ms)",
        "P50 AVG (ms)", "P99 AVG (ms)", "P99.9 AVG (ms)",
        "AQM Drops", "Congestion Sanctioned",
        "Packets", "Octets", "Max Rate (bps)",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 4, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    row = 5
    for sid, sf in sorted(service_flows.items()):
        bins = sf["latency_bins"]
        dir_str = "DS" if sf["direction"] == "downstream" else "US"
        edges = DS_BIN_EDGES_MS if dir_str == "DS" else US_BIN_EDGES_MS
        s = calc_latency_stats(bins, edges)

        _styled_cell(ws, row, 1, sid)
        _styled_cell(ws, row, 2, dir_str)
        _styled_cell(ws, row, 3, sf.get("service_class", ""))
        _styled_cell(ws, row, 4, s["total"], fill=_CALC_FILL)
        _styled_cell(ws, row, 5, round(s["weighted_avg"], 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 6, round(s["p50"],  4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 7, round(s["p99"],  4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 8, round(s["p999"], 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 9, round(s["p50a"],  4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 10, round(s["p99a"],  4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 11, round(s["p999a"], 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 12, sf["aqm_dropped_pkts"], fill=_CALC_FILL)
        _styled_cell(ws, row, 13, sf["congestion_sanctioned"], fill=_CALC_FILL)
        _styled_cell(ws, row, 14, sf["packets"], fill=_CALC_FILL)
        _styled_cell(ws, row, 15, sf["octets"], fill=_CALC_FILL)
        _styled_cell(ws, row, 16, sf.get("max_traffic_rate", 0), fill=_CALC_FILL)
        row += 1

    for i, w in enumerate([8, 12, 20, 14, 18, 12, 12, 12, 14, 14, 14, 12, 20, 12, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_throughput_sheet(wb, service_flows):
    ws = wb.create_sheet(title="Throughput")
    ws.sheet_properties.tabColor = "00B050"

    ws.merge_cells("A1:G1")
    ws["A1"] = "QOS SERVICE FLOW STATS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    headers = ["SID", "Direction", "Service Class", "Packets", "Octets",
               "AQM Drops", "Policed Drop Pkts"]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 3, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    row = 4
    total_pkts = 0
    total_octets = 0
    for sid, sf in sorted(service_flows.items()):
        dir_str = "DS" if sf["direction"] == "downstream" else "US"
        fill = _RESULT_FILL if sf["octets"] > 0 else None
        _styled_cell(ws, row, 1, sid)
        _styled_cell(ws, row, 2, dir_str, fill=fill)
        _styled_cell(ws, row, 3, sf.get("service_class", ""), fill=fill)
        _styled_cell(ws, row, 4, sf["packets"], fill=fill)
        _styled_cell(ws, row, 5, sf["octets"], fill=fill)
        _styled_cell(ws, row, 6, sf["aqm_dropped_pkts"], fill=fill)
        _styled_cell(ws, row, 7, sf["policed_drop_pkts"], fill=fill)
        total_pkts += sf["packets"]
        total_octets += sf["octets"]
        row += 1

    _styled_cell(ws, row, 1, "TOTAL", font=_BOLD, fill=_RESULT_FILL)
    _styled_cell(ws, row, 4, total_pkts, font=_BOLD, fill=_RESULT_FILL)
    _styled_cell(ws, row, 5, total_octets, font=_BOLD, fill=_RESULT_FILL)

    for i, w in enumerate([8, 12, 20, 14, 14, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------

def _print_sf_header(sf, sfid, dir_str):
    """Print the provisioned QoS header block for a service flow."""
    rate_mbps   = sf.get("max_traffic_rate", 0) / 1_000_000 if sf.get("max_traffic_rate") else 0
    burst_bytes = sf.get("max_traffic_burst", 0)
    concat_bytes= sf.get("max_concat_burst", 0)
    priority    = sf.get("priority", 0)
    svc_class   = sf.get("service_class", "")
    primary     = sf.get("primary", "")

    print(f"  {'─'*76}")
    print(f"  SID {sfid}  [{dir_str}]  {svc_class}  ({primary})")
    print(f"  {'─'*76}")
    print(f"  PROVISIONED QoS")
    print(f"    {'Max Rate':<18}: {rate_mbps:>10.0f} Mbps    {'Priority':<18}: {priority:>10}")
    print(f"    {'Max Burst':<18}: {burst_bytes:>10,} bytes   {'Max Concat Burst':<18}: {concat_bytes:>10,} bytes")


def print_summary(data, mac):
    """Print a formatted summary of all service flows to the console."""
    service_flows = extract_service_flow_data(data)
    if not service_flows:
        print(f"  No service flow data for {mac}")
        return

    item      = data["data"][0] if data.get("data") else {}
    firmware  = item.get("firmware_current_version", "N/A")
    modem_ip  = item.get("modem_ip", "N/A")
    reg_state = item.get("modem_reg_state", "N/A")
    docsis    = item.get("docsis_base_capability", "N/A")
    init      = item.get("init_state", "N/A")

    W = 80
    print()
    print("=" * W)
    print(f"  MODEM SUMMARY  —  {mac.upper()}")
    print("=" * W)
    print(f"  {'IP':<14}: {modem_ip:<20}  {'DOCSIS':<8}: {docsis}")
    print(f"  {'Reg State':<14}: {reg_state:<20}  {'Init':<8}: {init}")
    print(f"  {'Firmware':<14}: {firmware}")
    print(f"  {'Collected':<14}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    for sid, sf in sorted(service_flows.items()):
        dir_str   = "DS" if sf["direction"] == "downstream" else "US"
        edges     = DS_BIN_EDGES_MS if dir_str == "DS" else US_BIN_EDGES_MS
        bins      = sf["latency_bins"]
        s         = calc_latency_stats(bins, edges)
        has_data  = sf["packets"] > 0 or sf["octets"] > 0 or s["total"] > 0

        print()
        _print_sf_header(sf, sid, dir_str)

        if not has_data:
            print(f"    No traffic data")
            continue

        # Flow stats
        print(f"  {'FLOW STATS':}")
        print(f"    {'Packets':<18}: {sf['packets']:>18,}    {'Octets':<18}: {sf['octets']:>20,}")
        print(f"    {'AQM Drops':<18}: {sf['aqm_dropped_pkts']:>18,}    {'Policed Drops':<18}: {sf['policed_drop_pkts']:>20,}")
        print(f"    {'ECT1 Pkts':<18}: {sf['congestion_ect1']:>18,}    {'CE Marked':<18}: {sf['congestion_ce_marked']:>20,}")

        # Latency
        if s["total"] > 0:
            print()
            print(f"  LATENCY  ({s['total']:,} bin pkts)")
            print(f"    {'Metric':<20}  {'Interpolation':>16}    {'AVG Method':>16}")
            print(f"    {'-'*58}")
            print(f"    {'Weighted Avg':<20}  {s['weighted_avg']:>13.4f} ms")
            print(f"    {'P50':<20}  {s['p50']:>13.4f} ms    {s['p50a']:>13.4f} ms")
            print(f"    {'P99':<20}  {s['p99']:>13.4f} ms    {s['p99a']:>13.4f} ms")
            print(f"    {'P99.9':<20}  {s['p999']:>13.4f} ms    {s['p999a']:>13.4f} ms")
            print(f"    {'-'*58}")
            print(f"    {'Max Latency':<20}  {sf.get('max_latency', 0):>13,} ms    {'Hist Updates':<14}: {sf.get('hist_updates', 0):,}")
        else:
            print(f"\n  LATENCY  : no bin data")

    print()
    print("=" * W)
    print()


# ---------------------------------------------------------------------------
# Main report generation
# ---------------------------------------------------------------------------

def generate_report(data, mac, output_dir=None):
    """Generate Excel latency report from Scope API response data."""
    if not OPENPYXL_AVAILABLE:
        print("WARNING: openpyxl not available — skipping report generation")
        return None

    service_flows = extract_service_flow_data(data)
    if not service_flows:
        print(f"  No service flow data to report for {mac}")
        return None

    # Check if any latency bins have data
    has_bins = any(sum(sf["latency_bins"]) > 0 for sf in service_flows.values())
    if not has_bins:
        print(f"  NOTE: No latency bin data — report will contain throughput/congestion only")

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Reports")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"Scope_Modem_Report_{mac}_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Write both US and DS bin edge sheets so per-SF sheets reference the correct one
    write_bin_edges_sheet(wb, US_BIN_EDGES_MS, title="Bin_Edges_US", direction="US")
    write_bin_edges_sheet(wb, DS_BIN_EDGES_MS, title="Bin_Edges_DS", direction="DS")

    # Summary
    ts_display = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_summary_sheet(wb, service_flows, US_BIN_EDGES_MS, mac, ts_display)

    # Throughput
    write_throughput_sheet(wb, service_flows)

    # Per-service-flow sheets (only if bins have data)
    for sid, sf in sorted(service_flows.items()):
        if sum(sf["latency_bins"]) > 0:
            direction = "DS" if sf["direction"] == "downstream" else "US"
            edges = DS_BIN_EDGES_MS if direction == "DS" else US_BIN_EDGES_MS
            write_sf_sheet(wb, f"SID_{sid}", sf, edges, direction)

    wb.save(output_file)
    print(f"  Report saved: {output_file}")
    return output_file


# ---------------------------------------------------------------------------
# Main script
# ---------------------------------------------------------------------------
import sys
import time
import argparse

FIELDS = ",".join([
  "firmware_current_version", "modem_ip", "modem_mac", "modem_reg_state",
  "init_state", "docsis_base_capability",
  "qos_service_flow_direction", "qos_service_flow_primary", "qos_param_set_priority",
  "qos_param_set_max_traffic_rate", "qos_param_set_max_traffic_burst", "qos_param_set_max_concat_burst",
  "qos_service_us_stats_tx_retries", "qos_service_us_stats_tx_exceededs",
  "qos_service_us_stats_rq_retries", "qos_service_us_stats_rq_exceededs",
  "qos_param_set_service_class_name", "qos_service_flow_pkts", "qos_service_flow_octets",
  "qos_service_flow_policed_drop_kts", "qos_service_flow_policed_delay_pkts",
  "qos_service_flow_aqm_dropped_pkts", "qos_service_flow_sid",
  "qos_service_flow_buffer_size",
  "docsQosSfLatencyMaxLatency", "docsQosSfLatencyNumHistUpdates",
  "docsQosSfLatencyBin1Pkts", "docsQosSfLatencyBin2Pkts", "docsQosSfLatencyBin3Pkts",
  "docsQosSfLatencyBin4Pkts", "docsQosSfLatencyBin5Pkts", "docsQosSfLatencyBin6Pkts",
  "docsQosSfLatencyBin7Pkts", "docsQosSfLatencyBin8Pkts", "docsQosSfLatencyBin9Pkts",
  "docsQosSfLatencyBin10Pkts", "docsQosSfLatencyBin11Pkts", "docsQosSfLatencyBin12Pkts",
  "docsQosSfLatencyBin13Pkts", "docsQosSfLatencyBin14Pkts", "docsQosSfLatencyBin15Pkts",
  "docsQosSfLatencyBin16Pkts",
  "docsQosSfCongestionSanctionedPkts", "docsQosSfCongestionTotalEct0Pkts",
  "docsQosSfCongestionTotalEct1Pkts", "docsQosSfCongestionCeMarkedEct1Pkts",
  "docsQosSfCongestionArrivedCePkts"
])


def get_token():
    base_url = os.getenv('SCOPE_API_URL', 'https://scope-api.charter.com')
    user = os.getenv('SCOPE_API_USER')
    password = os.getenv('SCOPE_API_PASS')
    if not user or not password:
        print("ERROR: SCOPE_API_USER and SCOPE_API_PASS must be set in .env")
        sys.exit(1)
    url = f"{base_url}/v7/auth/login"
    payload = json.dumps({"username": user, "password": password, "clientId": "api"})
    response = requests.post(url, headers={"Content-Type": "application/json"}, data=payload)
    if response.status_code != 200:
        print(f"Login failed ({response.status_code}): {response.text}")
        sys.exit(1)
    resp_json = response.json()
    if "data" not in resp_json or "token" not in resp_json.get("data", {}):
        print(f"Login failed — unexpected response: {resp_json}")
        sys.exit(1)
    return f"Bearer {resp_json['data']['token']}"


def fetch_metrics(mac, token):
    url = f"https://scope-api.charter.com/v7/cablemodems/{mac}/metrics?fields={FIELDS}"
    headers = {"accept": "application/json", "authorization": token, "clientId": "web"}
    response = requests.get(url, headers=headers)
    data = response.json()
    if response.status_code != 200 or "error" in data or not data.get("data"):
        print(f"  ERROR ({response.status_code}): {data.get('message', 'No data available')}")
        return None
    return data


def compute_api_deltas(before_flows, after_flows):
    """Compute per-SFID bin and flow stat deltas between two API snapshots."""
    deltas = {}
    for sfid in sorted(set(before_flows) & set(after_flows)):
        b = before_flows[sfid]
        a = after_flows[sfid]
        bin_deltas = [max(a["latency_bins"][i] - b["latency_bins"][i], 0) for i in range(NUM_BINS)]
        d_packets = max(a["packets"] - b["packets"], 0)
        d_octets  = max(a["octets"]  - b["octets"],  0)
        d_aqm     = max(a["aqm_dropped_pkts"]   - b["aqm_dropped_pkts"],   0)
        d_policed = max(a["policed_drop_pkts"]   - b["policed_drop_pkts"],  0)
        d_ect1    = max(a["congestion_ect1"]      - b["congestion_ect1"],    0)
        d_ce      = max(a["congestion_ce_marked"] - b["congestion_ce_marked"], 0)
        deltas[sfid] = {
            "sf_key":             a["sf_key"],
            "direction":          a["direction"],
            "primary":            a["primary"],
            "service_class":      a["service_class"],
            "max_traffic_rate":   a["max_traffic_rate"],
            "max_latency":        a["max_latency"],
            "hist_updates":       a["hist_updates"],
            "latency_bins":       bin_deltas,
            "packets":            d_packets,
            "octets":             d_octets,
            "aqm_dropped_pkts":   d_aqm,
            "policed_drop_pkts":  d_policed,
            "policed_delay_pkts": max(a["policed_delay_pkts"] - b["policed_delay_pkts"], 0),
            "congestion_ect1":    d_ect1,
            "congestion_ce_marked": d_ce,
            "congestion_sanctioned": max(a["congestion_sanctioned"] - b["congestion_sanctioned"], 0),
            "congestion_ect0":    max(a["congestion_ect0"] - b["congestion_ect0"], 0),
            "congestion_arrived_ce": max(a["congestion_arrived_ce"] - b["congestion_arrived_ce"], 0),
            "tx_retries":   max(a["tx_retries"]   - b["tx_retries"],   0),
            "tx_exceededs": max(a["tx_exceededs"] - b["tx_exceededs"], 0),
            "rq_retries":   max(a["rq_retries"]   - b["rq_retries"],   0),
            "rq_exceededs": max(a["rq_exceededs"] - b["rq_exceededs"], 0),
            "buffer_size":  a["buffer_size"],
        }
    return deltas


def print_delta_summary(before_data, after_data, mac, wait_s):
    """Print before/after delta summary to console."""
    before_flows = extract_service_flow_data(before_data)
    after_flows  = extract_service_flow_data(after_data)
    delta_flows  = compute_api_deltas(before_flows, after_flows)

    item      = after_data["data"][0] if after_data.get("data") else {}
    firmware  = item.get("firmware_current_version", "N/A")
    modem_ip  = item.get("modem_ip", "N/A")
    reg_state = item.get("modem_reg_state", "N/A")
    docsis    = item.get("docsis_base_capability", "N/A")
    init      = item.get("init_state", "N/A")

    W = 80
    print()
    print("=" * W)
    print(f"  MODEM DELTA SUMMARY  —  {mac.upper()}  (wait={wait_s}s)")
    print("=" * W)
    print(f"  {'IP':<14}: {modem_ip:<20}  {'DOCSIS':<8}: {docsis}")
    print(f"  {'Reg State':<14}: {reg_state:<20}  {'Init':<8}: {init}")
    print(f"  {'Firmware':<14}: {firmware}")
    print(f"  {'Collected':<14}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    for sfid, sf in sorted(delta_flows.items()):
        dir_str   = "DS" if sf["direction"] == "downstream" else "US"
        edges     = DS_BIN_EDGES_MS if dir_str == "DS" else US_BIN_EDGES_MS
        s         = calc_latency_stats(sf["latency_bins"], edges)
        has_data  = sf["packets"] > 0 or sf["octets"] > 0 or s["total"] > 0

        print()
        _print_sf_header(sf, sfid, dir_str)

        if not has_data:
            print(f"    No traffic delta")
            continue

        # Throughput from octet delta
        throughput_mbps = (sf["octets"] * 8) / (wait_s * 1_000_000) if wait_s > 0 and sf["octets"] > 0 else 0

        print(f"  FLOW STATS DELTA  (over {wait_s}s)")
        print(f"    {'Packets':<18}: {sf['packets']:>18,}    {'Octets':<18}: {sf['octets']:>20,}")
        print(f"    {'Throughput':<18}: {throughput_mbps:>17.4f} Mbps")
        print(f"    {'AQM Drops':<18}: {sf['aqm_dropped_pkts']:>18,}    {'Policed Drops':<18}: {sf['policed_drop_pkts']:>20,}")
        print(f"    {'ECT1 Pkts':<18}: {sf['congestion_ect1']:>18,}    {'CE Marked':<18}: {sf['congestion_ce_marked']:>20,}")

        if s["total"] > 0:
            print()
            print(f"  LATENCY DELTA  ({s['total']:,} bin pkts)")
            print(f"    {'Metric':<20}  {'Interpolation':>16}    {'AVG Method':>16}")
            print(f"    {'-'*58}")
            print(f"    {'Weighted Avg':<20}  {s['weighted_avg']:>13.4f} ms")
            print(f"    {'P50':<20}  {s['p50']:>13.4f} ms    {s['p50a']:>13.4f} ms")
            print(f"    {'P99':<20}  {s['p99']:>13.4f} ms    {s['p99a']:>13.4f} ms")
            print(f"    {'P99.9':<20}  {s['p999']:>13.4f} ms    {s['p999a']:>13.4f} ms")
            print(f"    {'-'*58}")
            print(f"    {'Max Latency':<20}  {sf.get('max_latency', 0):>13,} ms    {'Hist Updates':<14}: {sf.get('hist_updates', 0):,}")
        else:
            print(f"\n  LATENCY  : no bin data")

    print()
    print("=" * W)
    print()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scope API modem metrics collector")
    parser.add_argument("--mac", action="append", dest="macs", metavar="MAC",
                        help="Modem MAC address (repeat for multiple)")
    parser.add_argument("--wait", type=int, default=0, metavar="SECONDS",
                        help="Seconds to wait between before/after collections (0 = single snapshot)")
    parser.add_argument("--interval", type=int, default=0, metavar="SECONDS",
                        help="Continuously poll every N seconds, printing delta each cycle (Ctrl+C to stop)")
    args = parser.parse_args()

    modem_macs = args.macs if args.macs else ["08a7c0885eff"]
    wait_s     = args.wait
    interval_s = args.interval

    token = get_token()
    print(f"Authenticated. Collecting metrics for: {', '.join(modem_macs)}")

    if interval_s > 0:
        # --- Continuous polling mode ---
        print(f"Polling every {interval_s}s — press Ctrl+C to stop\n")
        cycle = 0
        prev_data = {}
        try:
            while True:
                cycle += 1
                ts = datetime.now().strftime('%H:%M:%S')
                print(f"\n[Cycle {cycle}] {ts}")
                for mac in modem_macs:
                    current = fetch_metrics(mac, token)
                    if not current:
                        continue
                    if mac in prev_data:
                        print_delta_summary(prev_data[mac], current, mac, interval_s)
                    else:
                        print_summary(current, mac)
                        print(f"  {mac} — first snapshot collected, waiting for next cycle...")
                    prev_data[mac] = current
                for remaining in range(interval_s, 0, -1):
                    print(f"  {remaining}s remaining...", end="\r")
                    time.sleep(1)
                print()
        except KeyboardInterrupt:
            print(f"\nStopped after {cycle} cycle(s).")

    elif wait_s > 0:
        # --- Before snapshot ---
        print(f"\n[BEFORE] Collecting at {datetime.now().strftime('%H:%M:%S')}...")
        before_data = {}
        for mac in modem_macs:
            print(f"  {mac}")
            data = fetch_metrics(mac, token)
            if data:
                before_data[mac] = data
                print(json.dumps(data, indent=2))

        # --- Wait ---
        print(f"\nWaiting {wait_s} seconds...")
        for remaining in range(wait_s, 0, -1):
            print(f"  {remaining}s remaining...", end="\r")
            time.sleep(1)
        print()

        # --- After snapshot ---
        print(f"[AFTER] Collecting at {datetime.now().strftime('%H:%M:%S')}...")
        for mac in modem_macs:
            print(f"  {mac}")
            after = fetch_metrics(mac, token)
            if after and mac in before_data:
                print(json.dumps(after, indent=2))
                print_delta_summary(before_data[mac], after, mac, wait_s)
            elif after:
                print_summary(after, mac)
    else:
        # --- Single snapshot ---
        for mac in modem_macs:
            print(f"\n{'='*60}\n{mac}\n{'='*60}")
            data = fetch_metrics(mac, token)
            if data:
                print(json.dumps(data, indent=2))
                print_summary(data, mac)
